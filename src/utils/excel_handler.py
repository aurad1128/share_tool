# -*- coding: utf-8 -*-
"""
Excel 文件生成器
"""

import os
import subprocess
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from utils.storage import get_save_dir, add_history


# 样式常量
HEADER_FILL = PatternFill(start_color='FFE0687A', end_color='FFE0687A', fill_type='solid')
HEADER_FONT = Font(name='Microsoft YaHei', bold=True, color='FFFFFF', size=12)
DATA_FONT = Font(name='Microsoft YaHei', size=11)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
LEFT_ALIGN = Alignment(horizontal='left', vertical='center')
THIN_BORDER = Border(
    left=Side(style='thin', color='F0B0B8'),
    right=Side(style='thin', color='F0B0B8'),
    top=Side(style='thin', color='F0B0B8'),
    bottom=Side(style='thin', color='F0B0B8'),
)
ALT_FILL = PatternFill(start_color='FFFFF0F3', end_color='FFFFF0F3', fill_type='solid')


def save_to_excel(data, platform):
    """
    将热搜数据保存为 Excel 文件
    返回文件完整路径
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f'{platform}热搜'

    # 表头
    headers = ['排名', '标题', '热度值', '平台', '爬取时间']
    crawl_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    # 数据行
    for row_idx, item in enumerate(data, 2):
        values = [
            item.get('rank', ''),
            item.get('title', ''),
            item.get('heat', ''),
            platform,
            crawl_time,
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if col_idx == 2:  # 标题列左对齐
                cell.alignment = LEFT_ALIGN
            else:
                cell.alignment = CENTER_ALIGN
            # 交替行背景
            if row_idx % 2 == 0:
                cell.fill = ALT_FILL

    # 自适应列宽
    col_widths = {
        1: 8,    # 排名
        2: 45,   # 标题
        3: 18,   # 热度值
        4: 12,   # 平台
        5: 22,   # 爬取时间
    }
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 冻结表头
    ws.freeze_panes = 'A2'

    # 保存文件
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'{platform}热搜_{timestamp}.xlsx'
    save_dir = get_save_dir()
    filepath = os.path.join(save_dir, filename)
    wb.save(filepath)

    # 记录历史
    add_history({
        'platform': platform,
        'filename': filename,
        'filepath': filepath,
        'time': crawl_time,
        'count': len(data),
    })

    # 打开文件夹并选中文件
    _open_in_explorer(filepath)

    return filepath


def _open_in_explorer(filepath):
    """打开文件夹并选中文件"""
    try:
        subprocess.Popen(['explorer', '/select,', filepath])
    except Exception:
        pass
